import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from io import BytesIO
import datetime
import os

TEMPLATE_PATH = "template.xlsx"

# Load Excel template
@st.cache_data
def load_template():
    if not os.path.exists(TEMPLATE_PATH):
        st.error("Template file not found. Please ensure 'template.xlsx' exists in the working directory.")
        st.stop()
    return load_workbook(TEMPLATE_PATH)

# Save to BytesIO for download
def save_to_bytes(wb):
    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)
    return excel_io

st.set_page_config(page_title="GCC HR Reporting Dashboard", layout="wide")
st.markdown("""
    <style>
        .block-container {
            padding-top: 2rem;
            padding-bottom: 2rem;
        }
    </style>
""", unsafe_allow_html=True)

st.title("📊 GCC HR Reporting Dashboard")

# 1. Executive Summary
st.header("1. Executive Summary")
report_period = st.text_input("Report Period", "Q2 2025")
reporting_date = st.date_input("Reporting Date", value=datetime.date.today())
total_headcount = st.number_input("Total Headcount", min_value=0)
total_attrition = st.number_input("Total Attrition Rate (YTD)", min_value=0.0, step=0.1)
key_highlights = st.text_area("Key Highlights (one per line)")

# 2. Headcount Overview
st.header("2. Headcount Overview")
females = st.number_input("Female Employees", min_value=0)
contractors = st.number_input("Contractors", min_value=0)
interns = st.number_input("Interns/Trainees", min_value=0)
location_bangalore = st.number_input("Bangalore", min_value=0)
location_pune = st.number_input("Pune", min_value=0)
location_remote = st.number_input("Remote", min_value=0)

# 3. Skill Mix
st.header("3. Skill Mix")
skill_data = {}
for skill in ["Data Engineering", "Data Science / ML", "DevOps / SRE", "Cybersecurity", "Business Analytics", "Product/Program Mgmt", "Others"]:
    skill_data[skill] = st.number_input(f"{skill}", min_value=0)

# 4. Role Mix
st.header("4. Role Mix")
role_data = {}
for role in ["Individual Contributors", "Team Leads / SMEs", "Mid-level Managers", "Senior Leadership", "Support Functions (HR, IT, Finance)"]:
    role_data[role] = st.number_input(f"{role}", min_value=0)

# 5. Tenure Mix
st.header("5. Tenure Mix")
tenure_data = {}
for band in ["< 1 year", "1–3 years", "3–5 years", "5+ years"]:
    tenure_data[band] = st.number_input(f"{band}", min_value=0)

# 6. Attrition Analysis
st.header("6. Attrition Analysis")
vol_attr = st.number_input("Voluntary Attrition", min_value=0)
inv_attr = st.number_input("Involuntary Attrition", min_value=0)
exits_interviewed = st.radio("Exit Interviews Completed?", ["Yes", "No"])
exit_reasons = st.text_area("Top Reasons for Exit")
backfill_rate = st.number_input("Backfill Rate % (within 30 days)", min_value=0.0, step=0.1)

# 7. D&I
st.header("7. Diversity & Inclusion")
gender_div = st.number_input("Gender Diversity (% Female)", min_value=0.0, step=0.1)
avg_age = st.number_input("Average Age", min_value=0.0, step=0.1)
returning_mothers = st.number_input("% Returning Mothers", min_value=0.0, step=0.1)
lgbtq = st.number_input("LGBTQ+ Inclusion (Self-declared %)", min_value=0.0, step=0.1)
disability = st.number_input("Disability Inclusion %", min_value=0.0, step=0.1)

# 8. Hiring Overview
st.header("8. Hiring Overview")
new_hires = st.number_input("New Hires", min_value=0)
promotions = st.number_input("Internal Movements / Promotions", min_value=0)
lateral_pct = st.number_input("% Lateral Hires", min_value=0.0, step=0.1)
offer_to_join_ratio = st.number_input("Offer-to-Join Ratio %", min_value=0.0, step=0.1)
avg_time_to_fill = st.number_input("Average Time to Fill (in Days)", min_value=0.0, step=1.0)

# 9. L&D
st.header("9. Learning & Development")
avg_training = st.number_input("Avg. Training Hours per Employee", min_value=0.0, step=0.1)
trained_pct = st.number_input("% Employees Trained in Last 6 Months", min_value=0.0, step=0.1)
certifications = st.number_input("Certifications Completed", min_value=0)
mgr_dev = st.number_input("Managerial Development Coverage %", min_value=0.0, step=0.1)

# 10. Engagement
st.header("10. Employee Engagement")
engagement_index = st.number_input("Engagement Index %", min_value=0.0, step=0.1)
enps = st.number_input("eNPS", min_value=-100, max_value=100, step=1)
survey_response_rate = st.number_input("Survey Response Rate %", min_value=0.0, step=0.1)
recognitions = st.number_input("Recognition Awards Issued", min_value=0)

# 11. Compliance
st.header("11. Compliance & Policy Metrics")
posh_training = st.radio("POSH Committee & Training Coverage", ["Yes", "No"])
bgv = st.radio("Background Verifications Completed", ["Yes", "No"])
mandatory_trainings = st.radio("Mandatory Trainings Completed", ["Yes", "No"])
grievances = st.number_input("Grievances Resolved", min_value=0)

# Generate Excel
if st.button("📄 Generate Report"):
    wb = load_template()
    ws1 = wb["Executive Summary"]
    ws1["B2"] = report_period
    ws1["B3"] = str(reporting_date)
    ws1["B4"] = total_headcount
    ws1["B5"] = total_attrition
    ws1["B6"] = key_highlights

    # Populate other sheets
    wb["Headcount Overview"]["B2"] = total_headcount
    wb["Headcount Overview"]["B3"] = females
    wb["Headcount Overview"]["B4"] = contractors
    wb["Headcount Overview"]["B5"] = interns
    wb["Headcount Overview"]["B6"] = location_bangalore
    wb["Headcount Overview"]["B7"] = location_pune
    wb["Headcount Overview"]["B8"] = location_remote

    for idx, (skill, count) in enumerate(skill_data.items(), start=2):
        wb["Skill Mix"].cell(row=idx, column=2, value=count)

    for idx, (role, count) in enumerate(role_data.items(), start=2):
        wb["Role Mix"].cell(row=idx, column=2, value=count)

    for idx, (band, count) in enumerate(tenure_data.items(), start=2):
        wb["Tenure Mix"].cell(row=idx, column=2, value=count)

    wb["Attrition Analysis"]["B2"] = vol_attr
    wb["Attrition Analysis"]["B3"] = inv_attr
    wb["Attrition Analysis"]["B4"] = exits_interviewed
    wb["Attrition Analysis"]["B5"] = exit_reasons
    wb["Attrition Analysis"]["B6"] = backfill_rate

    wb["Diversity & Inclusion"]["B2"] = gender_div
    wb["Diversity & Inclusion"]["B3"] = avg_age
    wb["Diversity & Inclusion"]["B4"] = returning_mothers
    wb["Diversity & Inclusion"]["B5"] = lgbtq
    wb["Diversity & Inclusion"]["B6"] = disability

    wb["Hiring Overview"]["B2"] = new_hires
    wb["Hiring Overview"]["B3"] = promotions
    wb["Hiring Overview"]["B4"] = lateral_pct
    wb["Hiring Overview"]["B5"] = offer_to_join_ratio
    wb["Hiring Overview"]["B6"] = avg_time_to_fill

    wb["Learning & Development"]["B2"] = avg_training
    wb["Learning & Development"]["B3"] = trained_pct
    wb["Learning & Development"]["B4"] = certifications
    wb["Learning & Development"]["B5"] = mgr_dev

    wb["Employee Engagement"]["B2"] = engagement_index
    wb["Employee Engagement"]["B3"] = enps
    wb["Employee Engagement"]["B4"] = survey_response_rate
    wb["Employee Engagement"]["B5"] = recognitions

    wb["Compliance & Policy Metrics"]["B2"] = posh_training
    wb["Compliance & Policy Metrics"]["B3"] = bgv
    wb["Compliance & Policy Metrics"]["B4"] = mandatory_trainings
    wb["Compliance & Policy Metrics"]["B5"] = grievances

    output = save_to_bytes(wb)
    st.success("✅ Report ready!")
    st.download_button("📥 Download HR Report", data=output, file_name="GCC_HR_Report.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
